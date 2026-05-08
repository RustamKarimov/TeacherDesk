from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
import re
import time

from django.conf import settings
from django.db import OperationalError
from django.db import close_old_connections
from openpyxl import load_workbook
from pypdf import PdfReader
from pypdf import PdfWriter


@dataclass(frozen=True)
class ParsedExamCode:
    subject_code: str
    session: str
    document_type: str
    component: str
    paper_number: int


def _retry_locked_database(operation, *, attempts: int = 6, delay: float = 0.25):
    for attempt in range(attempts):
        try:
            return operation()
        except OperationalError as error:
            if "database is locked" not in str(error).lower() or attempt == attempts - 1:
                raise
            close_old_connections()
            time.sleep(delay * (attempt + 1))
    raise RuntimeError("Database operation did not complete.")


def parse_cambridge_exam_code(exam_code: str) -> ParsedExamCode:
    match = re.fullmatch(r"(?P<subject>\d+)_(?P<session>[a-z]\d{2})_(?P<doc>qp|ms)_(?P<component>\d+)", exam_code)
    if not match:
        raise ValueError(f"Exam code does not match Cambridge pattern: {exam_code}")

    component = match.group("component")
    return ParsedExamCode(
        subject_code=match.group("subject"),
        session=match.group("session"),
        document_type=match.group("doc"),
        component=component,
        paper_number=int(component[0]),
    )


def parse_starred_page(value: object) -> tuple[int, bool]:
    text = str(value).strip()
    is_starred = text.endswith("*")
    number_text = text[:-1] if is_starred else text
    if not number_text.isdigit():
        raise ValueError(f"Invalid page value: {value}")
    return int(number_text), is_starred


def question_bank_paths(subject_code: str, paper_number: int, question_number: int, qp_filename: str, ms_filename: str) -> dict[str, str]:
    base = f"QuestionBank/{subject_code}/Paper{paper_number}/Q{question_number}"
    return {
        "question_folder": f"{base}/Questions",
        "markscheme_folder": f"{base}/MarkSchemes",
        "question_file": f"{base}/Questions/{qp_filename}",
        "markscheme_file": f"{base}/MarkSchemes/{ms_filename}",
    }


REQUIRED_MANIFEST_COLUMNS = [
    "exam_code",
    "question_number",
    "qp_start_page",
    "ms_start_page",
    "Mark",
    "T1",
    "T2",
    "T3",
    "topic_1",
    "topic_2",
    "topic_3",
]


def _pdf_page_count(path: Path, cache: dict[str, int]) -> int | None:
    key = str(path)
    if key in cache:
        return cache[key]
    if not path.exists():
        return None
    reader = PdfReader(str(path))
    cache[key] = len(reader.pages)
    return cache[key]


def _severity_counts(items: list[dict[str, object]]) -> dict[str, int]:
    counts = {"error": 0, "warning": 0, "info": 0}
    for item in items:
        severity = str(item.get("severity", "info"))
        counts[severity] = counts.get(severity, 0) + 1
    return counts


def validate_manifest(manifest_path: str, source_root: str) -> dict[str, object]:
    manifest = Path(manifest_path)
    source = Path(source_root)
    issues: list[dict[str, object]] = []
    rows: list[dict[str, object]] = []
    topics: dict[int, str] = {}
    pdf_cache: dict[str, int] = {}

    if not manifest.exists():
        return {
            "ok": False,
            "summary": {"rows": 0, "errors": 1, "warnings": 0, "starred_boundaries": 0, "review_required_items": 0},
            "issues": [
                {
                    "severity": "error",
                    "row": None,
                    "exam_code": None,
                    "message": f"Manifest file was not found: {manifest}",
                }
            ],
            "rows": [],
            "topics": [],
        }

    workbook = load_workbook(manifest, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    raw_headers = next(sheet.iter_rows(values_only=True), None)
    headers = [str(value).strip() if value is not None else "" for value in raw_headers or []]
    missing_columns = [column for column in REQUIRED_MANIFEST_COLUMNS if column not in headers]
    if missing_columns:
        issues.append(
            {
                "severity": "error",
                "row": 1,
                "exam_code": None,
                "message": f"Missing required columns: {', '.join(missing_columns)}",
            }
        )

    if len(workbook.sheetnames) > 1:
        topic_sheet = workbook[workbook.sheetnames[1]]
        for topic_row in topic_sheet.iter_rows(min_row=2, values_only=True):
            number, name = (list(topic_row) + [None, None])[:2]
            if isinstance(number, int) and name:
                topics[number] = str(name).strip()

    header_index = {name: index for index, name in enumerate(headers)}
    by_exam: dict[str, list[dict[str, object]]] = defaultdict(list)

    for excel_row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None for value in values):
            continue

        def get(column: str) -> object:
            index = header_index.get(column)
            return values[index] if index is not None and index < len(values) else None

        exam_code = str(get("exam_code") or "").strip()
        question_number = get("question_number")
        row_issues: list[dict[str, object]] = []

        try:
            parsed = parse_cambridge_exam_code(exam_code)
        except ValueError as exc:
            row_issues.append({"severity": "error", "message": str(exc)})
            parsed = None

        try:
            qp_start_page, qp_starred = parse_starred_page(get("qp_start_page"))
        except ValueError as exc:
            row_issues.append({"severity": "error", "message": str(exc)})
            qp_start_page, qp_starred = None, False

        try:
            ms_start_page, ms_starred = parse_starred_page(get("ms_start_page"))
        except ValueError as exc:
            row_issues.append({"severity": "error", "message": str(exc)})
            ms_start_page, ms_starred = None, False

        try:
            question_number_int = int(question_number)
        except (TypeError, ValueError):
            row_issues.append({"severity": "error", "message": f"Invalid question_number: {question_number}"})
            question_number_int = None

        try:
            marks = int(get("Mark"))
        except (TypeError, ValueError):
            row_issues.append({"severity": "warning", "message": f"Invalid or missing Mark value: {get('Mark')}"})
            marks = None

        qp_path = source / f"{exam_code}.pdf"
        ms_exam_code = exam_code.replace("_qp_", "_ms_")
        ms_path = source / f"{ms_exam_code}.pdf"

        qp_pages = None
        ms_pages = None
        if parsed:
            if not qp_path.exists():
                row_issues.append({"severity": "error", "message": f"Question paper file not found: {qp_path}"})
            else:
                try:
                    qp_pages = _pdf_page_count(qp_path, pdf_cache)
                except Exception as exc:  # noqa: BLE001 - the exact parser failure is useful to the user.
                    row_issues.append({"severity": "error", "message": f"Could not read question paper PDF: {exc}"})

            if not ms_path.exists():
                row_issues.append({"severity": "error", "message": f"Mark scheme file not found: {ms_path}"})
            else:
                try:
                    ms_pages = _pdf_page_count(ms_path, pdf_cache)
                except Exception as exc:  # noqa: BLE001
                    row_issues.append({"severity": "error", "message": f"Could not read mark scheme PDF: {exc}"})

        if qp_pages and qp_start_page and qp_start_page > qp_pages:
            row_issues.append({"severity": "error", "message": f"qp_start_page {qp_start_page} is beyond QP page count {qp_pages}"})
        if ms_pages and ms_start_page and ms_start_page > ms_pages:
            row_issues.append({"severity": "error", "message": f"ms_start_page {ms_start_page} is beyond MS page count {ms_pages}"})

        assigned_topics = []
        for topic_number_column, topic_name_column in (("T1", "topic_1"), ("T2", "topic_2"), ("T3", "topic_3")):
            topic_number = get(topic_number_column)
            topic_name = get(topic_name_column)
            if topic_name:
                assigned_topics.append(str(topic_name).strip())
            if isinstance(topic_number, int) and topic_name and topics.get(topic_number) and topics[topic_number] != str(topic_name).strip():
                row_issues.append(
                    {
                        "severity": "warning",
                        "message": f"{topic_number_column}={topic_number} maps to '{topics[topic_number]}', but row has '{topic_name}'",
                    }
                )

        if not assigned_topics:
            row_issues.append({"severity": "info", "message": "No topics assigned. This is allowed."})

        record = {
            "row": excel_row_number,
            "exam_code": exam_code,
            "question_number": question_number_int,
            "paper": f"Paper{parsed.paper_number}" if parsed else None,
            "component": parsed.component if parsed else None,
            "marks": marks,
            "topics": assigned_topics,
            "qp_start_page": qp_start_page,
            "ms_start_page": ms_start_page,
            "qp_starred": qp_starred,
            "ms_starred": ms_starred,
            "qp_file_exists": qp_path.exists(),
            "ms_file_exists": ms_path.exists(),
            "qp_pages": qp_pages,
            "ms_pages": ms_pages,
            "issues": row_issues,
        }
        rows.append(record)
        if exam_code:
            by_exam[exam_code].append(record)

        for row_issue in row_issues:
            issues.append(
                {
                    "severity": row_issue["severity"],
                    "row": excel_row_number,
                    "exam_code": exam_code,
                    "question_number": question_number_int,
                    "message": row_issue["message"],
                }
            )

    review_required: set[tuple[str, int, str]] = set()
    starred_boundaries = 0
    for exam_code, exam_rows in by_exam.items():
        sorted_rows = sorted(exam_rows, key=lambda item: int(item["question_number"] or 0))
        seen_questions: dict[int, int] = {}
        for row in sorted_rows:
            question_number = row.get("question_number")
            if not isinstance(question_number, int):
                continue
            previous_row = seen_questions.get(question_number)
            if previous_row is not None:
                issues.append(
                    {
                        "severity": "error",
                        "row": row["row"],
                        "exam_code": exam_code,
                        "question_number": question_number,
                        "message": (
                            f"Duplicate question_number Q{question_number} in {exam_code}. "
                            f"Rows {previous_row} and {row['row']} both describe the same question. "
                            "Fix the duplicate row in the manifest, then run Check & preview again."
                        ),
                    }
                )
            seen_questions[question_number] = int(row["row"])

        for start_key, starred_key, label, column_name in (
            ("qp_start_page", "qp_starred", "Question paper", "qp_start_page"),
            ("ms_start_page", "ms_starred", "Mark scheme", "ms_start_page"),
        ):
            for index, row in enumerate(sorted_rows[:-1]):
                next_row = sorted_rows[index + 1]
                current_start = row.get(start_key)
                next_start = next_row.get(start_key)
                current_question = row.get("question_number")
                next_question = next_row.get("question_number")
                if not isinstance(current_start, int) or not isinstance(next_start, int):
                    continue
                if next_start < current_start:
                    issues.append(
                        {
                            "severity": "error",
                            "row": next_row["row"],
                            "exam_code": exam_code,
                            "question_number": next_question,
                            "message": (
                                f"{label} page starts go backwards in {exam_code}: row {next_row['row']} "
                                f"Q{next_question} has {column_name}={next_start}, but previous row {row['row']} "
                                f"Q{current_question} has {column_name}={current_start}. "
                                f"Fix {column_name} for Q{next_question} in the manifest, then run Check & preview again."
                            ),
                        }
                    )
                elif next_start == current_start and not row.get(starred_key):
                    issues.append(
                        {
                            "severity": "error",
                            "row": next_row["row"],
                            "exam_code": exam_code,
                            "question_number": next_question,
                            "message": (
                                f"{label} page boundary is shared in {exam_code}: row {row['row']} Q{current_question} "
                                f"and row {next_row['row']} Q{next_question} both have {column_name}={next_start}. "
                                f"If both questions really begin on this same page, change row {row['row']} "
                                f"Q{current_question} {column_name} to {current_start}* so both split PDFs are marked for review. "
                                f"If Q{next_question} starts later, correct row {next_row['row']} {column_name}. "
                                "Then run Check & preview again."
                            ),
                        }
                    )

        for index, row in enumerate(sorted_rows):
            next_row = sorted_rows[index + 1] if index + 1 < len(sorted_rows) else None
            if row["qp_starred"]:
                starred_boundaries += 1
                review_required.add((exam_code, int(row["question_number"]), "QP"))
                if next_row:
                    review_required.add((exam_code, int(next_row["question_number"]), "QP"))
            if row["ms_starred"]:
                starred_boundaries += 1
                review_required.add((exam_code, int(row["question_number"]), "MS"))
                if next_row:
                    review_required.add((exam_code, int(next_row["question_number"]), "MS"))

    counts = _severity_counts(issues)
    return {
        "ok": counts["error"] == 0,
        "summary": {
            "rows": len(rows),
            "errors": counts["error"],
            "warnings": counts["warning"],
            "info": counts["info"],
            "starred_boundaries": starred_boundaries,
            "review_required_items": len(review_required),
            "topics": len(topics),
            "source_root": str(source),
            "manifest_path": str(manifest),
        },
        "issues": issues[:250],
        "rows": rows[:100],
        "topics": [{"number": number, "name": name} for number, name in sorted(topics.items())],
    }


def _iter_manifest_records(manifest_path: str) -> tuple[list[dict[str, object]], dict[int, str]]:
    workbook = load_workbook(manifest_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    raw_headers = next(sheet.iter_rows(values_only=True), None)
    headers = [str(value).strip() if value is not None else "" for value in raw_headers or []]
    header_index = {name: index for index, name in enumerate(headers)}

    topics: dict[int, str] = {}
    if len(workbook.sheetnames) > 1:
        topic_sheet = workbook[workbook.sheetnames[1]]
        for topic_row in topic_sheet.iter_rows(min_row=2, values_only=True):
            number, name = (list(topic_row) + [None, None])[:2]
            if isinstance(number, int) and name:
                topics[number] = str(name).strip()

    records = []
    for excel_row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None for value in values):
            continue

        def get(column: str) -> object:
            index = header_index.get(column)
            return values[index] if index is not None and index < len(values) else None

        exam_code = str(get("exam_code") or "").strip()
        parsed = parse_cambridge_exam_code(exam_code)
        qp_start_page, qp_starred = parse_starred_page(get("qp_start_page"))
        ms_start_page, ms_starred = parse_starred_page(get("ms_start_page"))

        topic_values = []
        for topic_number_column, topic_name_column in (("T1", "topic_1"), ("T2", "topic_2"), ("T3", "topic_3")):
            topic_number = get(topic_number_column)
            topic_name = get(topic_name_column)
            if topic_name:
                topic_values.append(
                    {
                        "number": int(topic_number) if isinstance(topic_number, int) else None,
                        "name": str(topic_name).strip(),
                    }
                )

        records.append(
            {
                "row": excel_row_number,
                "exam_code": exam_code,
                "ms_exam_code": exam_code.replace("_qp_", "_ms_"),
                "question_number": int(get("question_number")),
                "marks": int(get("Mark")) if get("Mark") is not None else None,
                "subject_code": parsed.subject_code,
                "session": parsed.session,
                "component": parsed.component,
                "paper_number": parsed.paper_number,
                "qp_start_page": qp_start_page,
                "qp_start_page_raw": str(get("qp_start_page")),
                "qp_starred": qp_starred,
                "ms_start_page": ms_start_page,
                "ms_start_page_raw": str(get("ms_start_page")),
                "ms_starred": ms_starred,
                "topics": topic_values,
            }
        )

    return records, topics


def _range_for_row(rows: list[dict[str, object]], index: int, start_key: str, starred_key: str, total_pages: int) -> tuple[int, int]:
    current = rows[index]
    start = int(current[start_key])
    next_row = rows[index + 1] if index + 1 < len(rows) else None
    if next_row is None:
        end = total_pages
    else:
        next_start = int(next_row[start_key])
        end = next_start if current[starred_key] else next_start - 1
    return start, min(end, total_pages)


def _starred_boundary_message(document_type: str, position: str, current_question: int, adjacent_question: int | None) -> str:
    label = "Question paper" if document_type == "QP" else "Mark scheme"
    if adjacent_question is None:
        return f"{label} needs review: starred page boundary found. Review the shared page and remove any overlapping content if needed."
    if position == "end":
        return f"{label} needs review: final page may include the beginning of Q{adjacent_question}. Review the split PDF and remove any overlapping content if needed."
    return f"{label} needs review: first page may include the end of Q{current_question}. Review the split PDF and remove any overlapping content if needed."


def build_starred_boundary_reviews(
    grouped_rows: dict[str, list[dict[str, object]]],
) -> tuple[set[tuple[str, int, str]], dict[tuple[str, int, str], list[str]]]:
    review_required: set[tuple[str, int, str]] = set()
    review_notes: dict[tuple[str, int, str], list[str]] = defaultdict(list)

    for exam_code, exam_rows in grouped_rows.items():
        exam_rows.sort(key=lambda item: int(item["question_number"]))
        for index, record in enumerate(exam_rows):
            question_number = int(record["question_number"])
            next_record = exam_rows[index + 1] if index + 1 < len(exam_rows) else None
            next_question_number = int(next_record["question_number"]) if next_record else None

            for starred_key, document_type in (("qp_starred", "QP"), ("ms_starred", "MS")):
                if not record[starred_key]:
                    continue

                current_key = (exam_code, question_number, document_type)
                review_required.add(current_key)
                review_notes[current_key].append(
                    _starred_boundary_message(document_type, "end", question_number, next_question_number)
                )

                if next_question_number is not None:
                    next_key = (exam_code, next_question_number, document_type)
                    review_required.add(next_key)
                    review_notes[next_key].append(
                        _starred_boundary_message(document_type, "start", question_number, next_question_number)
                    )

    return review_required, review_notes


def _write_pdf_range(source_path: Path, output_path: Path, start_page: int, end_page: int) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    reader = PdfReader(str(source_path))
    writer = PdfWriter()
    for page_number in range(start_page, end_page + 1):
        writer.add_page(reader.pages[page_number - 1])
    with output_path.open("wb") as output_file:
        writer.write(output_file)


def versioned_output_path(output_path: Path, version_number: int, start_page: int, end_page: int) -> Path:
    version_folder = output_path.parent / "_versions"
    stem = output_path.stem
    suffix = output_path.suffix
    return version_folder / f"{stem}__v{version_number}_pages_{start_page}-{end_page}{suffix}"


def next_version_number(output_path: Path) -> int:
    version_folder = output_path.parent / "_versions"
    if not version_folder.exists():
        return 2
    existing = list(version_folder.glob(f"{output_path.stem}__v*_pages_*{output_path.suffix}"))
    return len(existing) + 2


def import_and_split_manifest(
    manifest_path: str,
    source_root: str,
    *,
    output_root: str | None = None,
    library_name: str = "Cambridge Physics Library",
    overwrite: bool = False,
    existing_pdf_strategy: str = "skip",
    changed_page_strategy: str = "flag",
    metadata_strategy: str = "update",
    preserve_review_status: bool = True,
    dry_run: bool = False,
    progress_callback=None,
) -> dict[str, object]:
    validation = validate_manifest(manifest_path, source_root)
    if not validation["ok"]:
        return {
            "ok": False,
            "summary": {
                "created_questions": 0,
                "updated_questions": 0,
                "split_question_pdfs": 0,
                "split_markscheme_pdfs": 0,
                "skipped_existing_files": 0,
                "review_required_items": validation["summary"]["review_required_items"],
            },
            "message": "Import and split was stopped because validation has errors.",
            "validation": validation,
            "outputs": [],
        }

    records, topic_catalog = _iter_manifest_records(manifest_path)
    source = Path(source_root)
    library_root = Path(output_root).resolve() if output_root else settings.TEACHERDESK_LIBRARY_ROOT
    outputs: list[dict[str, object]] = []

    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        grouped[str(record["exam_code"])].append(record)
    review_required, review_notes = build_starred_boundary_reviews(grouped)

    if dry_run:
        return {
            "ok": True,
            "summary": {
                "created_questions": 0,
                "updated_questions": 0,
                "split_question_pdfs": 0,
                "split_markscheme_pdfs": 0,
                "skipped_existing_files": 0,
                "review_required_items": len(review_required),
                "records_ready": len(records),
                "library_root": str(library_root),
            },
            "message": "Dry run passed. No database records or PDFs were written.",
            "validation": validation,
            "outputs": [],
        }

    from apps.catalog.models import Question, Topic
    from apps.libraries.models import Library
    from apps.splitter.models import ManifestImport

    library, _ = _retry_locked_database(
        lambda: Library.objects.get_or_create(
            name=library_name,
            defaults={
                "root_path": str(library_root),
                "source_papers_path": str(source),
                "question_bank_path": "QuestionBank",
                "generated_exams_path": "generated_exams",
                "manifests_path": "manifests",
                "naming_preset": "cambridge",
            },
        )
    )

    _retry_locked_database(
        lambda: ManifestImport.objects.create(
            library=library,
            name=Path(manifest_path).name,
            file_path=manifest_path,
            status=ManifestImport.Status.IMPORTED,
            row_count=len(records),
            error_count=0,
            warning_count=int(validation["summary"]["warnings"]),
            report=validation,
        )
    )

    topic_objects: dict[tuple[str, int | None, str], Topic] = {}
    for number, name in topic_catalog.items():
        topic, _ = _retry_locked_database(
            lambda number=number, name=name: Topic.objects.get_or_create(
                subject_code="9702",
                topic_number=number,
                name=name,
                defaults={"source": "manifest"},
            )
        )
        topic_objects[("9702", number, name)] = topic

    created_questions = 0
    updated_questions = 0
    split_question_pdfs = 0
    split_markscheme_pdfs = 0
    skipped_existing_files = 0
    processed_files = 0
    total_files = len(records) * 2

    def report_progress(current: dict[str, object] | None = None) -> None:
        if progress_callback:
            progress_callback(
                {
                    "processed_files": processed_files,
                    "total_files": total_files,
                    "split_question_pdfs": split_question_pdfs,
                    "split_markscheme_pdfs": split_markscheme_pdfs,
                    "skipped_existing_files": skipped_existing_files,
                    "current": current or {},
                }
            )

    for exam_code, exam_rows in grouped.items():
        exam_rows.sort(key=lambda item: int(item["question_number"]))
        qp_path = source / f"{exam_code}.pdf"
        ms_path = source / f"{str(exam_rows[0]['ms_exam_code'])}.pdf"
        qp_total_pages = len(PdfReader(str(qp_path)).pages)
        ms_total_pages = len(PdfReader(str(ms_path)).pages)

        for index, record in enumerate(exam_rows):
            question_number = int(record["question_number"])
            subject_code = str(record["subject_code"])
            paper_number = int(record["paper_number"])
            qp_output_name = f"{exam_code}_Q{question_number}.pdf"
            ms_output_name = f"{record['ms_exam_code']}_Q{question_number}.pdf"
            paths = question_bank_paths(subject_code, paper_number, question_number, qp_output_name, ms_output_name)
            qp_output = library_root / paths["question_file"]
            ms_output = library_root / paths["markscheme_file"]

            qp_start, qp_end = _range_for_row(exam_rows, index, "qp_start_page", "qp_starred", qp_total_pages)
            ms_start, ms_end = _range_for_row(exam_rows, index, "ms_start_page", "ms_starred", ms_total_pages)

            existing_question = Question.objects.filter(library=library, exam_code=exam_code, question_number=question_number).first()
            qp_range_changed = bool(
                existing_question
                and existing_question.qp_page_start is not None
                and existing_question.qp_page_end is not None
                and (existing_question.qp_page_start, existing_question.qp_page_end) != (qp_start, qp_end)
            )
            ms_range_changed = bool(
                existing_question
                and existing_question.ms_page_start is not None
                and existing_question.ms_page_end is not None
                and (existing_question.ms_page_start, existing_question.ms_page_end) != (ms_start, ms_end)
            )

            qp_write_path = qp_output
            if qp_output.exists() and qp_range_changed and changed_page_strategy == "keep_both":
                qp_write_path = versioned_output_path(qp_output, next_version_number(qp_output), qp_start, qp_end)
            should_write_qp = (
                overwrite
                or not qp_output.exists()
                or existing_pdf_strategy == "overwrite"
                or (qp_range_changed and changed_page_strategy in {"overwrite", "keep_both"})
            )
            if should_write_qp:
                _write_pdf_range(qp_path, qp_write_path, qp_start, qp_end)
                split_question_pdfs += 1
            else:
                skipped_existing_files += 1
            processed_files += 1
            report_progress({"exam_code": exam_code, "question_number": question_number, "document_type": "QP", "output": str(qp_write_path)})

            ms_write_path = ms_output
            if ms_output.exists() and ms_range_changed and changed_page_strategy == "keep_both":
                ms_write_path = versioned_output_path(ms_output, next_version_number(ms_output), ms_start, ms_end)
            should_write_ms = (
                overwrite
                or not ms_output.exists()
                or existing_pdf_strategy == "overwrite"
                or (ms_range_changed and changed_page_strategy in {"overwrite", "keep_both"})
            )
            if should_write_ms:
                _write_pdf_range(ms_path, ms_write_path, ms_start, ms_end)
                split_markscheme_pdfs += 1
            else:
                skipped_existing_files += 1
            processed_files += 1
            report_progress({"exam_code": exam_code, "question_number": question_number, "document_type": "MS", "output": str(ms_write_path)})

            qp_review = Question.ReviewStatus.NEEDS_REVIEW if (exam_code, question_number, "QP") in review_required else Question.ReviewStatus.NOT_REQUIRED
            ms_review = Question.ReviewStatus.NEEDS_REVIEW if (exam_code, question_number, "MS") in review_required else Question.ReviewStatus.NOT_REQUIRED
            qp_reasons = list(review_notes.get((exam_code, question_number, "QP"), []))
            ms_reasons = list(review_notes.get((exam_code, question_number, "MS"), []))
            if qp_range_changed and changed_page_strategy == "flag":
                qp_review = Question.ReviewStatus.NEEDS_REVIEW
                qp_reasons.append("Question paper needs review: page range changed; existing PDF was kept.")
            if ms_range_changed and changed_page_strategy == "flag":
                ms_review = Question.ReviewStatus.NEEDS_REVIEW
                ms_reasons.append("Mark scheme needs review: page range changed; existing PDF was kept.")
            if qp_range_changed and changed_page_strategy == "keep_both":
                qp_review = Question.ReviewStatus.NEEDS_REVIEW
                qp_reasons.append("Question paper needs review: page range changed; new version was saved in _versions.")
            if ms_range_changed and changed_page_strategy == "keep_both":
                ms_review = Question.ReviewStatus.NEEDS_REVIEW
                ms_reasons.append("Mark scheme needs review: page range changed; new version was saved in _versions.")

            qp_db_path = qp_output
            qp_db_start = qp_start
            qp_db_end = qp_end
            if existing_question and changed_page_strategy == "keep_both" and qp_range_changed:
                qp_db_path = Path(existing_question.split_qp_path) if existing_question.split_qp_path else qp_output
                qp_db_start = existing_question.qp_page_start or qp_start
                qp_db_end = existing_question.qp_page_end or qp_end

            ms_db_path = ms_output
            ms_db_start = ms_start
            ms_db_end = ms_end
            if existing_question and changed_page_strategy == "keep_both" and ms_range_changed:
                ms_db_path = Path(existing_question.split_ms_path) if existing_question.split_ms_path else ms_output
                ms_db_start = existing_question.ms_page_start or ms_start
                ms_db_end = existing_question.ms_page_end or ms_end

            defaults = {
                    "subject_code": subject_code,
                    "session": str(record["session"]),
                    "component": str(record["component"]),
                    "paper_number": paper_number,
                    "marks": record["marks"],
                    "source_qp_path": str(qp_path),
                    "source_ms_path": str(ms_path),
                    "split_qp_path": str(qp_db_path),
                    "split_ms_path": str(ms_db_path),
                    "qp_start_page_raw": str(record["qp_start_page_raw"]),
                    "ms_start_page_raw": str(record["ms_start_page_raw"]),
                    "qp_page_start": qp_db_start,
                    "qp_page_end": qp_db_end,
                    "ms_page_start": ms_db_start,
                    "ms_page_end": ms_db_end,
                    "qp_review_status": qp_review,
                    "ms_review_status": ms_review,
            }
            if existing_question and preserve_review_status:
                if existing_question.qp_review_status == Question.ReviewStatus.REVIEWED:
                    defaults["qp_review_status"] = existing_question.qp_review_status
                if existing_question.ms_review_status == Question.ReviewStatus.REVIEWED:
                    defaults["ms_review_status"] = existing_question.ms_review_status

            reasons = []
            if defaults["qp_review_status"] == Question.ReviewStatus.NEEDS_REVIEW:
                reasons.extend(qp_reasons)
            if defaults["ms_review_status"] == Question.ReviewStatus.NEEDS_REVIEW:
                reasons.extend(ms_reasons)
            defaults["review_reason"] = " ".join(reasons)

            if metadata_strategy == "keep" and existing_question:
                question = existing_question
                created = False
            else:
                question, created = _retry_locked_database(
                    lambda: Question.objects.update_or_create(
                        library=library,
                        exam_code=exam_code,
                        question_number=question_number,
                        defaults=defaults,
                    )
                )
            if created:
                created_questions += 1
            else:
                updated_questions += 1

            row_topics = []
            for topic_value in record["topics"]:
                topic_key = (subject_code, topic_value["number"], topic_value["name"])
                topic = topic_objects.get(topic_key)
                if topic is None:
                    topic, _ = _retry_locked_database(
                        lambda topic_value=topic_value: Topic.objects.get_or_create(
                            subject_code=subject_code,
                            topic_number=topic_value["number"],
                            name=topic_value["name"],
                            defaults={"source": "manifest"},
                        )
                    )
                    topic_objects[topic_key] = topic
                row_topics.append(topic)
            if metadata_strategy != "keep" or created:
                _retry_locked_database(lambda question=question, row_topics=row_topics: question.topics.set(row_topics))

            if len(outputs) < 50:
                outputs.append(
                    {
                        "exam_code": exam_code,
                        "question_number": question_number,
                        "qp_pages": [qp_start, qp_end],
                        "ms_pages": [ms_start, ms_end],
                        "qp_output": str(qp_output),
                        "ms_output": str(ms_output),
                        "qp_review_status": qp_review,
                        "ms_review_status": ms_review,
                    }
                )

    return {
        "ok": True,
        "summary": {
            "created_questions": created_questions,
            "updated_questions": updated_questions,
            "split_question_pdfs": split_question_pdfs,
            "split_markscheme_pdfs": split_markscheme_pdfs,
            "skipped_existing_files": skipped_existing_files,
            "review_required_items": len(review_required),
            "library_root": str(library_root),
        },
        "message": "Manifest imported and PDFs split successfully.",
        "validation": validation,
        "outputs": outputs,
    }


def build_split_plan(
    manifest_path: str,
    source_root: str,
    *,
    output_root: str | None = None,
    existing_pdf_strategy: str = "skip",
    changed_page_strategy: str = "flag",
    metadata_strategy: str = "update",
) -> dict[str, object]:
    validation = validate_manifest(manifest_path, source_root)
    library_root = Path(output_root).resolve() if output_root else settings.TEACHERDESK_LIBRARY_ROOT
    if not validation["ok"]:
        return {
            "ok": False,
            "message": "Split plan cannot be built until validation errors are fixed.",
            "summary": {
                "records": 0,
                "files_total": 0,
                "files_to_create": 0,
                "files_to_skip_existing": 0,
                "questions_to_create": 0,
                "questions_to_update": 0,
                "review_required_items": validation["summary"]["review_required_items"],
            },
            "validation": validation,
            "items": [],
        }

    records, _topic_catalog = _iter_manifest_records(manifest_path)
    items = []
    files_to_create = 0
    files_to_skip_existing = 0
    files_to_overwrite = 0
    files_to_version = 0
    page_range_changes = 0
    questions_to_create = 0
    questions_to_update = 0
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)

    for record in records:
        grouped[str(record["exam_code"])].append(record)
    review_required, _review_notes = build_starred_boundary_reviews(grouped)

    existing_questions: dict[tuple[str, int], object] = {}
    try:
        from apps.catalog.models import Question
        from apps.libraries.models import Library

        library = Library.objects.filter(root_path=str(library_root)).first()
        if library:
            existing_questions = {(question.exam_code, question.question_number): question for question in Question.objects.filter(library=library)}
    except Exception:  # noqa: BLE001 - planning should still work before DB is fully ready.
        existing_questions = {}

    for record in records:
        exam_code = str(record["exam_code"])
        question_number = int(record["question_number"])
        qp_output_name = f"{exam_code}_Q{question_number}.pdf"
        ms_output_name = f"{record['ms_exam_code']}_Q{question_number}.pdf"
        paths = question_bank_paths(str(record["subject_code"]), int(record["paper_number"]), question_number, qp_output_name, ms_output_name)
        qp_output = library_root / paths["question_file"]
        ms_output = library_root / paths["markscheme_file"]
        qp_exists = qp_output.exists()
        ms_exists = ms_output.exists()
        existing_question = existing_questions.get((exam_code, question_number))
        qp_range_changed = bool(existing_question and existing_question.qp_page_start and (existing_question.qp_page_start, existing_question.qp_page_end) != (None, None))
        ms_range_changed = bool(existing_question and existing_question.ms_page_start and (existing_question.ms_page_start, existing_question.ms_page_end) != (None, None))
        # We calculate exact ranges by opening the PDFs only after the broad output plan is known.
        if existing_question:
            source = Path(source_root)
            grouped_rows = grouped[exam_code]
            record_index = grouped_rows.index(record)
            qp_total_pages = len(PdfReader(str(source / f"{exam_code}.pdf")).pages)
            ms_total_pages = len(PdfReader(str(source / f"{record['ms_exam_code']}.pdf")).pages)
            qp_start, qp_end = _range_for_row(grouped_rows, record_index, "qp_start_page", "qp_starred", qp_total_pages)
            ms_start, ms_end = _range_for_row(grouped_rows, record_index, "ms_start_page", "ms_starred", ms_total_pages)
            qp_range_changed = (
                existing_question.qp_page_start is not None
                and existing_question.qp_page_end is not None
                and (existing_question.qp_page_start, existing_question.qp_page_end) != (qp_start, qp_end)
            )
            ms_range_changed = (
                existing_question.ms_page_start is not None
                and existing_question.ms_page_end is not None
                and (existing_question.ms_page_start, existing_question.ms_page_end) != (ms_start, ms_end)
            )

        for exists, changed in ((qp_exists, qp_range_changed), (ms_exists, ms_range_changed)):
            if not exists:
                files_to_create += 1
            elif changed and changed_page_strategy == "overwrite":
                files_to_overwrite += 1
            elif changed and changed_page_strategy == "keep_both":
                files_to_version += 1
            elif existing_pdf_strategy == "overwrite":
                files_to_overwrite += 1
            else:
                files_to_skip_existing += 1
        page_range_changes += int(qp_range_changed) + int(ms_range_changed)

        if (exam_code, question_number) in existing_questions:
            questions_to_update += 1
        else:
            questions_to_create += 1

        if len(items) < 100:
            items.append(
                {
                    "exam_code": exam_code,
                    "question_number": question_number,
                    "paper": f"Paper{record['paper_number']}",
                    "qp_output": str(qp_output),
                    "ms_output": str(ms_output),
                    "qp_action": _planned_file_action(qp_exists, qp_range_changed, existing_pdf_strategy, changed_page_strategy),
                    "ms_action": _planned_file_action(ms_exists, ms_range_changed, existing_pdf_strategy, changed_page_strategy),
                    "qp_range_changed": qp_range_changed,
                    "ms_range_changed": ms_range_changed,
                    "qp_review_required": (exam_code, question_number, "QP") in review_required,
                    "ms_review_required": (exam_code, question_number, "MS") in review_required,
                }
            )

    return {
        "ok": True,
        "message": (
            "Split plan built. Existing PDFs will be overwritten."
            if existing_pdf_strategy == "overwrite"
            else "Split plan built. Existing PDFs will be skipped."
        ),
        "summary": {
            "records": len(records),
            "files_total": len(records) * 2,
            "files_to_create": files_to_create,
            "files_to_skip_existing": files_to_skip_existing,
            "files_to_overwrite": files_to_overwrite,
            "files_to_version": files_to_version,
            "page_range_changes": page_range_changes,
            "questions_to_create": questions_to_create,
            "questions_to_update": questions_to_update if metadata_strategy == "update" else 0,
            "review_required_items": len(review_required),
            "library_root": str(library_root),
        },
        "validation": validation,
        "items": items,
    }


def _planned_file_action(exists: bool, changed: bool, existing_pdf_strategy: str, changed_page_strategy: str) -> str:
    if not exists:
        return "create"
    if changed and changed_page_strategy == "overwrite":
        return "overwrite_changed"
    if changed and changed_page_strategy == "keep_both":
        return "save_version"
    if existing_pdf_strategy == "overwrite":
        return "overwrite"
    if changed:
        return "skip_flag_review"
    return "skip_existing"
